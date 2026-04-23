"""Скрипт для анализа статистики нажатий на кнопки бота."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

LOGS_DIR = Path(__file__).resolve().parent / "logs"
USER_ACTIVITY_LOG = LOGS_DIR / "user_activity.log"

BUTTON_CATEGORIES = {
    "main_menu": ["start"],
    "business_open": ["how_open_business"],
    "non_financial": ["non_fin_support", "non_fin_mb"],
    "financial": ["fin_support", "fin_mb", "fin_garant"],
    "property": ["property_support", "prop_kovorking", "prop_peregovornaya", "prop_conf_zal", "prop_maly_zal"],
    "contacts": ["contacts_orgs", "contacts_mb"],
    "callback": ["callback_request", "callback_consult", "callback_platform"],
    "evaluate": ["evaluate_quality", "evaluate_mb"],
    "chatbot": ["chat_bot_info", "chat_exit_to_menu"],
    "other_services": ["productivity_labor", "agro_support", "export_coop", "education_services"],
}

CATEGORY_NAMES = {
    "main_menu": "Главное меню",
    "business_open": "Открытие бизнеса",
    "non_financial": "Не финансовая поддержка",
    "financial": "Финансовая поддержка",
    "property": "Имущество",
    "contacts": "Контакты",
    "callback": "Звонок",
    "evaluate": "Оценка услуг",
    "chatbot": "Чат-бот",
    "other_services": "Другие услуги",
    "navigation": "Навигация (страницы)",
    "other": "Прочее",
}

BUTTON_NAMES = {
    "start": "Начать",
    "back_to_main": "В главное меню",
    "back_main": "В главное меню",
    "how_open_business": "Открыть бизнес",
    "non_fin_support": "Нефинансовая поддержка",
    "non_fin_mb": "Центр «Мой бизнес» (нефин.)",
    "fin_support": "Финансовая поддержка",
    "fin_mb": "Мой бизнес (займы)",
    "fin_garant": "Гарантийный фонд",
    "property_support": "Имущественная поддержка",
    "prop_kovorking": "Аренда коворкинга",
    "prop_peregovornaya": "Аренда переговорной",
    "prop_conf_zal": "Аренда конференц-зала",
    "prop_maly_zal": "Аренда малого зала",
    "productivity_labor": "Производительность труда",
    "agro_support": "Поддержка АПК",
    "export_coop": "Экспорт",
    "education_services": "Обучение",
    "contacts_orgs": "Контакты",
    "callback_request": "Обратный звонок",
    "callback_consult": "Консультация «Мой бизнес»",
    "callback_platform": "Проблемы с МСП.РФ",
    "evaluate_quality": "Оценить услуги",
    "evaluate_mb": "Оценить Центр «Мой бизнес»",
    "chat_bot_info": "Чат-бот",
    "chat_exit_to_menu": "Выйти из диалога",
}


def parse_log_line(line: str) -> dict[str, Any] | None:
    """Парсит строку лога и извлекает данные о callback."""
    pattern = r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) \| first_name='([^']*)' last_name='([^']*)' user_id=(\S+) chat_id=(\S+) action=(\S+)$"
    match = re.match(pattern, line.strip())
    if not match:
        return None

    timestamp_str, first_name, last_name, user_id, chat_id, action = match.groups()

    if not action.startswith("callback:"):
        return None

    button = action.removeprefix("callback:")

    return {
        "timestamp": datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S"),
        "first_name": first_name,
        "last_name": last_name,
        "user_id": user_id,
        "chat_id": chat_id,
        "button": button,
    }


def get_button_category(button: str) -> str:
    """Определяет категорию кнопки."""
    if re.match(r"^(non_fin_page_|fin_mb_page_|fin_mb_details_|fin_mb_open_)", button):
        return "navigation"

    # Навигационные кнопки "назад" - исключаем из основной статистики
    navigation_buttons = [
        "back_to_main", "back_main", "back_non_fin_org", "back_non_fin_services",
        "back_fin_org", "back_property_services", "back_contacts_orgs",
        "back_callback_menu", "back_evaluate_menu"
    ]
    if button in navigation_buttons:
        return "navigation"

    for category, buttons in BUTTON_CATEGORIES.items():
        if button in buttons:
            return category

    return "other"


def get_button_name(button: str) -> str:
    """Возвращает человекочитаемое название кнопки."""
    if button in BUTTON_NAMES:
        return BUTTON_NAMES[button]

    if match := re.match(r"non_fin_page_(\d+)", button):
        return f"Нефин. поддержка - страница {int(match.group(1)) + 1}"
    if match := re.match(r"fin_mb_page_(\d+)", button):
        return f"Фин. поддержка - страница {int(match.group(1)) + 1}"
    if match := re.match(r"fin_mb_details_(\d+)", button):
        return f"Фин. поддержка - детали {int(match.group(1)) + 1}"
    if match := re.match(r"fin_mb_open_(\d+)", button):
        return f"Фин. поддержка - открыть {int(match.group(1)) + 1}"

    return button


def analyze_logs(days: int | None = None) -> dict[str, Any]:
    """Анализирует логи и возвращает статистику."""
    if not USER_ACTIVITY_LOG.exists():
        return {"error": f"Лог-файл не найден: {USER_ACTIVITY_LOG}"}

    entries: list[dict[str, Any]] = []
    cutoff_date = datetime.now() - timedelta(days=days) if days else None

    with open(USER_ACTIVITY_LOG, "r", encoding="utf-8") as f:
        for line in f:
            entry = parse_log_line(line)
            if entry:
                if cutoff_date and entry["timestamp"] < cutoff_date:
                    continue
                entries.append(entry)

    if not entries:
        return {"error": "Нет данных за указанный период"}

    total_clicks = len(entries)
    unique_users = len(set(e["user_id"] for e in entries))
    unique_chats = len(set(e["chat_id"] for e in entries))

    button_counter = Counter(e["button"] for e in entries)

    category_counter: Counter[str] = Counter()
    for button in button_counter:
        category = get_button_category(button)
        category_counter[category] += button_counter[button]

    user_clicks: defaultdict[str, int] = defaultdict(int)
    user_buttons: defaultdict[str, set[str]] = defaultdict(set)
    user_names: defaultdict[str, dict[str, str]] = defaultdict(lambda: {"first_name": "", "last_name": ""})
    for e in entries:
        user_clicks[e["user_id"]] += 1
        user_buttons[e["user_id"]].add(e["button"])
        user_names[e["user_id"]]["first_name"] = e["first_name"]
        user_names[e["user_id"]]["last_name"] = e["last_name"]

    daily_activity: defaultdict[str, int] = defaultdict(int)
    for e in entries:
        day = e["timestamp"].strftime("%Y-%m-%d")
        daily_activity[day] += 1

    hourly_activity: defaultdict[int, int] = defaultdict(int)
    for e in entries:
        hourly_activity[e["timestamp"].hour] += 1

    return {
        "period_days": days,
        "total_entries": len(entries),
        "total_clicks": total_clicks,
        "unique_users": unique_users,
        "unique_chats": unique_chats,
        "button_stats": dict(button_counter.most_common()),
        "category_stats": dict(category_counter.most_common()),
        "user_stats": {
            "most_active": [(user_id, clicks, user_names[user_id]["first_name"], user_names[user_id]["last_name"]) for user_id, clicks in Counter(user_clicks).most_common(10)],
            "avg_clicks_per_user": total_clicks / unique_users if unique_users > 0 else 0,
        },
        "daily_activity": dict(sorted(daily_activity.items())),
        "hourly_activity": dict(sorted(hourly_activity.items())),
        "date_range": {
            "from": min(e["timestamp"] for e in entries).strftime("%Y-%m-%d %H:%M:%S"),
            "to": max(e["timestamp"] for e in entries).strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


def save_csv_report(stats: dict[str, Any], output_path: Path) -> Path:
    """Сохраняет отчёт в CSV формате."""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        
        # Заголовок
        writer.writerow(["ОТЧЁТ ПО СТАТИСТИКЕ НАЖАТИЙ НА КНОПКИ"])
        writer.writerow(["Сформирован:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])
        
        # Общая статистика
        writer.writerow(["ОСНОВНАЯ СТАТИСТИКА"])
        writer.writerow(["Показатель", "Значение"])
        writer.writerow(["Период", f"{stats['date_range']['from']} - {stats['date_range']['to']}"])
        writer.writerow(["Всего нажатий", stats['total_clicks']])
        writer.writerow(["Уникальных пользователей", stats['unique_users']])
        writer.writerow(["Уникальных чатов", stats['unique_chats']])
        writer.writerow(["Среднее нажатий на пользователя", f"{stats['user_stats']['avg_clicks_per_user']:.1f}"])
        writer.writerow([])
        
        # Статистика по категориям
        writer.writerow(["СТАТИСТИКА ПО КАТЕГОРИЯМ"])
        writer.writerow(["Категория", "Количество", "Процент"])
        for category, count in stats['category_stats'].items():
            category_name = CATEGORY_NAMES.get(category, category)
            percentage = (count / stats['total_clicks']) * 100
            writer.writerow([category_name, count, f"{percentage:.1f}%"])
        writer.writerow([])
        
        # Топ кнопок
        writer.writerow(["ТОП-20 ПОПУЛЯРНЫХ КНОПОК"])
        writer.writerow(["Место", "Название кнопки", "Количество", "Процент"])
        for i, (button, count) in enumerate(list(stats['button_stats'].items())[:20], 1):
            button_name = get_button_name(button)
            percentage = (count / stats['total_clicks']) * 100
            writer.writerow([i, button_name, count, f"{percentage:.1f}%"])
        writer.writerow([])
        
        # Самые активные пользователи
        writer.writerow(["САМЫЕ АКТИВНЫЕ ПОЛЬЗОВАТЕЛИ"])
        writer.writerow(["Место", "ID пользователя", "Имя", "Фамилия", "Количество нажатий"])
        for i, (user_id, clicks, first_name, last_name) in enumerate(stats['user_stats']['most_active'], 1):
            writer.writerow([i, user_id, first_name, last_name, clicks])
        writer.writerow([])
        
        # Активность по часам
        writer.writerow(["АКТИВНОСТЬ ПО ЧАСАМ"])
        writer.writerow(["Час", "Количество"])
        for hour in range(24):
            if hour in stats['hourly_activity']:
                writer.writerow([f"{hour:02d}:00", stats['hourly_activity'][hour]])
    
    return output_path


def save_excel_report(stats: dict[str, Any], output_path: Path) -> Path:
    """Сохраняет отчёт в Excel формате с оформлением."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    
    # Стили
    header_fill = PatternFill(start_color="1a5490", end_color="1a5490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1a5490")
    subheader_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Лист 1: Общая статистика
    ws1 = wb.active
    ws1.title = "Общая статистика"
    
    ws1['A1'] = "Отчёт по статистике нажатий на кнопки"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    ws1['A2'] = "Сформирован:"
    ws1['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    row = 4
    ws1[f'A{row}'] = "ОСНОВНАЯ СТАТИСТИКА"
    ws1[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    headers = ["Показатель", "Значение"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    data = [
        ["Период", f"{stats['date_range']['from']} - {stats['date_range']['to']}"],
        ["Всего нажатий", stats['total_clicks']],
        ["Уникальных пользователей", stats['unique_users']],
        ["Уникальных чатов", stats['unique_chats']],
        ["Среднее нажатий на пользователя", f"{stats['user_stats']['avg_clicks_per_user']:.1f}"],
    ]
    
    for r in data:
        for col, val in enumerate(r, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
        row += 1
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 40
    
    # Лист 2: Статистика по категориям
    ws2 = wb.create_sheet("По категориям")
    
    ws2['A1'] = "Статистика по категориям"
    ws2['A1'].font = title_font
    
    row = 3
    headers = ["Категория", "Количество", "Процент"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for category, count in stats['category_stats'].items():
        category_name = CATEGORY_NAMES.get(category, category)
        percentage = (count / stats['total_clicks']) * 100
        ws2.cell(row=row, column=1, value=category_name).border = border
        ws2.cell(row=row, column=2, value=count).border = border
        ws2.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        row += 1
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    
    # Лист 3: Топ кнопок
    ws3 = wb.create_sheet("Топ кнопок")
    
    ws3['A1'] = "ТОП-20 ПОПУЛЯРНЫХ КНОПОК"
    ws3['A1'].font = title_font
    
    row = 3
    headers = ["Место", "Название кнопки", "Количество", "Процент"]
    header_fill_red = PatternFill(start_color="c62828", end_color="c62828", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=row, column=col, value=header)
        cell.fill = header_fill_red
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for i, (button, count) in enumerate(list(stats['button_stats'].items())[:20], 1):
        button_name = get_button_name(button)
        percentage = (count / stats['total_clicks']) * 100
        ws3.cell(row=row, column=1, value=i).border = border
        ws3.cell(row=row, column=2, value=button_name).border = border
        ws3.cell(row=row, column=3, value=count).border = border
        ws3.cell(row=row, column=4, value=f"{percentage:.1f}%").border = border
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        row += 1
    
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 15
    
    # Лист 4: Пользователи
    ws4 = wb.create_sheet("Пользователи")
    
    ws4['A1'] = "САМЫЕ АКТИВНЫЕ ПОЛЬЗОВАТЕЛИ"
    ws4['A1'].font = title_font
    
    row = 3
    headers = ["Место", "ID пользователя", "Имя", "Фамилия", "Количество нажатий"]
    header_fill_orange = PatternFill(start_color="f57c00", end_color="f57c00", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=row, column=col, value=header)
        cell.fill = header_fill_orange
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for i, (user_id, clicks, first_name, last_name) in enumerate(stats['user_stats']['most_active'], 1):
        ws4.cell(row=row, column=1, value=i).border = border
        ws4.cell(row=row, column=2, value=str(user_id)).border = border
        ws4.cell(row=row, column=3, value=first_name).border = border
        ws4.cell(row=row, column=4, value=last_name).border = border
        ws4.cell(row=row, column=5, value=clicks).border = border
        ws4.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws4.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws4.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        row += 1

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 20
    
    # Лист 5: Активность по часам
    ws5 = wb.create_sheet("По часам")

    ws5['A1'] = "Активность по часам"
    ws5['A1'].font = title_font

    peak_hour = max(stats['hourly_activity'].items(), key=lambda x: x[1])
    ws5['A2'] = f"Пиковый час: {peak_hour[0]:02d}:00 ({peak_hour[1]} нажатий)"

    row = 4
    headers = ["Час", "Количество"]
    header_fill_purple = PatternFill(start_color="6a1b9a", end_color="6a1b9a", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=row, column=col, value=header)
        cell.fill = header_fill_purple
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for hour in range(24):
        if hour in stats['hourly_activity']:
            count = stats['hourly_activity'][hour]
            ws5.cell(row=row, column=1, value=f"{hour:02d}:00").border = border
            ws5.cell(row=row, column=2, value=count).border = border
            ws5.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws5.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1

    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 15
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Анализ статистики нажатий на кнопки бота")
    parser.add_argument("--days", "-d", type=int, help="Анализировать за последние N дней")
    parser.add_argument("--format", "-f", choices=["csv", "excel", "json"], default="excel",
                        help="Формат вывода (csv, excel или json)")
    parser.add_argument("--output", "-o", type=str, help="Сохранить отчёт в файл")
    parser.add_argument("--no-save", action="store_true", help="Не сохранять отчёт в файл")

    args = parser.parse_args()

    stats = analyze_logs(days=args.days)
    
    if "error" in stats:
        print(f"Ошибка: {stats['error']}")
        return
    
    # Сохраняем отчёт
    if not args.no_save:
        if args.output:
            output_path = Path(args.output)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "xlsx" if args.format == "excel" else ("csv" if args.format == "csv" else "json")
            output_path = LOGS_DIR / f"button_stats_report_{timestamp}.{ext}"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if args.format == "json":
            output_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
        elif args.format == "csv":
            save_csv_report(stats, output_path)
        else:  # excel
            if not OPENPYXL_AVAILABLE:
                print("openpyxl не установлен. Сохраняю в CSV.")
                output_path = output_path.with_suffix(".csv")
                save_csv_report(stats, output_path)
            else:
                save_excel_report(stats, output_path)
        
        print(f"Отчёт сохранён: {output_path}")
    

if __name__ == "__main__":
    main()